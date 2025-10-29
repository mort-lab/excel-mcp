"""Cell operations"""

from openpyxl import load_workbook

from ..models import (
    CellReadRequest,
    CellResult,
    CellWriteRequest,
    RangeReadRequest,
    RangeResult,
    RangeWriteRequest,
)
from ..utils.validators import validate_file_path, validate_range_reference


def write_cell_value(request: CellWriteRequest) -> CellResult:
    """
    Write a value to a specific cell.

    Args:
        request: CellWriteRequest with workbook path, sheet name, cell, and value

    Returns:
        CellResult with success status
    """
    try:
        # Validate file path
        is_valid, error = validate_file_path(request.workbook_path, must_exist=True)
        if not is_valid:
            return CellResult(success=False, message=error)

        # Load workbook
        wb = load_workbook(request.workbook_path)

        # Check if sheet exists
        if request.sheet_name not in wb.sheetnames:
            wb.close()
            return CellResult(
                success=False,
                message=f"Sheet '{request.sheet_name}' not found. Available sheets: {wb.sheetnames}",
            )

        # Get worksheet
        ws = wb[request.sheet_name]

        # Write value
        ws[request.cell] = request.value

        # Save workbook
        wb.save(request.workbook_path)
        wb.close()

        return CellResult(
            success=True,
            message=f"Value written to {request.cell}",
            cell=request.cell,
            value=request.value,
        )

    except Exception as e:
        return CellResult(success=False, message=f"Failed to write cell: {str(e)}")


def read_cell_value(request: CellReadRequest) -> CellResult:
    """
    Read a value from a specific cell.

    Args:
        request: CellReadRequest with workbook path, sheet name, and cell

    Returns:
        CellResult with the cell value
    """
    try:
        # Validate file path
        is_valid, error = validate_file_path(request.workbook_path, must_exist=True)
        if not is_valid:
            return CellResult(success=False, message=error)

        # Load workbook
        wb = load_workbook(request.workbook_path, data_only=True)

        # Check if sheet exists
        if request.sheet_name not in wb.sheetnames:
            wb.close()
            return CellResult(
                success=False,
                message=f"Sheet '{request.sheet_name}' not found. Available sheets: {wb.sheetnames}",
            )

        # Get worksheet
        ws = wb[request.sheet_name]

        # Read value
        cell_obj = ws[request.cell]
        value = cell_obj.value

        wb.close()

        return CellResult(success=True, message=f"Value read from {request.cell}", cell=request.cell, value=value)

    except Exception as e:
        return CellResult(success=False, message=f"Failed to read cell: {str(e)}")


def write_range_values(request: RangeWriteRequest) -> RangeResult:
    """
    Write data to a range of cells.

    Args:
        request: RangeWriteRequest with workbook path, sheet name, start cell, and data

    Returns:
        RangeResult with success status
    """
    try:
        # Validate file path
        is_valid, error = validate_file_path(request.workbook_path, must_exist=True)
        if not is_valid:
            return RangeResult(success=False, message=error)

        # Validate data
        if not request.data or not request.data[0]:
            return RangeResult(success=False, message="Data cannot be empty")

        # Load workbook
        wb = load_workbook(request.workbook_path)

        # Check if sheet exists
        if request.sheet_name not in wb.sheetnames:
            wb.close()
            return RangeResult(
                success=False,
                message=f"Sheet '{request.sheet_name}' not found. Available sheets: {wb.sheetnames}",
            )

        # Get worksheet
        ws = wb[request.sheet_name]

        # Get starting cell coordinates
        start_cell_obj = ws[request.start_cell]
        start_row = start_cell_obj.row
        start_col = start_cell_obj.column

        # Write data
        rows_written = 0
        cols_written = 0

        for row_idx, row_data in enumerate(request.data):
            for col_idx, value in enumerate(row_data):
                ws.cell(row=start_row + row_idx, column=start_col + col_idx, value=value)
                cols_written = max(cols_written, col_idx + 1)
            rows_written += 1

        # Save workbook
        wb.save(request.workbook_path)
        wb.close()

        return RangeResult(
            success=True,
            message=f"Data written to range starting at {request.start_cell}",
            range=f"{request.start_cell}",
            rows=rows_written,
            cols=cols_written,
        )

    except Exception as e:
        return RangeResult(success=False, message=f"Failed to write range: {str(e)}")


def read_range_values(request: RangeReadRequest) -> RangeResult:
    """
    Read data from a range of cells.

    Args:
        request: RangeReadRequest with workbook path, sheet name, and range

    Returns:
        RangeResult with the data from the range
    """
    try:
        # Validate file path
        is_valid, error = validate_file_path(request.workbook_path, must_exist=True)
        if not is_valid:
            return RangeResult(success=False, message=error)

        # Validate range reference
        is_valid, error = validate_range_reference(request.range_ref)
        if not is_valid:
            return RangeResult(success=False, message=error)

        # Load workbook
        wb = load_workbook(request.workbook_path, data_only=True)

        # Check if sheet exists
        if request.sheet_name not in wb.sheetnames:
            wb.close()
            return RangeResult(
                success=False,
                message=f"Sheet '{request.sheet_name}' not found. Available sheets: {wb.sheetnames}",
            )

        # Get worksheet
        ws = wb[request.sheet_name]

        # Read range
        cell_range = ws[request.range_ref]

        # Convert to 2D list
        data = []
        if isinstance(cell_range, tuple):
            # Multiple rows
            for row in cell_range:
                if isinstance(row, tuple):
                    data.append([cell.value for cell in row])
                else:
                    data.append([row.value])
        else:
            # Single cell
            data = [[cell_range.value]]

        wb.close()

        rows = len(data)
        cols = len(data[0]) if data else 0

        return RangeResult(
            success=True,
            message=f"Data read from range {request.range_ref}",
            range=request.range_ref,
            rows=rows,
            cols=cols,
            data=data,
        )

    except Exception as e:
        return RangeResult(success=False, message=f"Failed to read range: {str(e)}")


def write_formula(workbook_path: str, sheet_name: str, cell: str, formula: str) -> CellResult:
    """
    Write a formula to a cell.

    Args:
        workbook_path: Path to the Excel workbook
        sheet_name: Name of the worksheet
        cell: Cell reference
        formula: Excel formula (should start with '=')

    Returns:
        CellResult with success status
    """
    try:
        # Validate file path
        is_valid, error = validate_file_path(workbook_path, must_exist=True)
        if not is_valid:
            return CellResult(success=False, message=error)

        # Ensure formula starts with =
        if not formula.startswith("="):
            formula = f"={formula}"

        # Load workbook
        wb = load_workbook(workbook_path)

        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            wb.close()
            return CellResult(
                success=False,
                message=f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}",
            )

        # Get worksheet
        ws = wb[sheet_name]

        # Write formula
        ws[cell] = formula

        # Save workbook
        wb.save(workbook_path)
        wb.close()

        return CellResult(success=True, message=f"Formula written to {cell}", cell=cell, value=formula)

    except Exception as e:
        return CellResult(success=False, message=f"Failed to write formula: {str(e)}")
