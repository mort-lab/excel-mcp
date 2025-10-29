"""Formatting operations"""

from typing import Any, cast

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..models import (
    AlignmentFormatRequest,
    BorderFormatRequest,
    FillFormatRequest,
    FontFormatRequest,
    NumberFormatRequest,
    OperationResult,
)
from ..utils.validators import validate_file_path, validate_range_reference


def format_font(request: FontFormatRequest) -> OperationResult:
    """
    Apply font formatting to a range of cells.

    Args:
        request: FontFormatRequest with formatting parameters

    Returns:
        OperationResult with success status
    """
    try:
        # Validate file path
        is_valid, error = validate_file_path(request.workbook_path, must_exist=True)
        if not is_valid:
            return OperationResult(success=False, message=error)

        # Validate range
        is_valid, error = validate_range_reference(request.range_ref)
        if not is_valid:
            return OperationResult(success=False, message=error)

        # Load workbook
        wb = load_workbook(request.workbook_path)

        # Check if sheet exists
        if request.sheet_name not in wb.sheetnames:
            wb.close()
            return OperationResult(
                success=False,
                message=f"Sheet '{request.sheet_name}' not found. Available sheets: {wb.sheetnames}",
            )

        ws = wb[request.sheet_name]

        # Create font object
        font_kwargs = {}
        if request.font_name:
            font_kwargs["name"] = request.font_name
        if request.font_size:
            font_kwargs["size"] = request.font_size
        if request.bold is not None:
            font_kwargs["bold"] = request.bold
        if request.italic is not None:
            font_kwargs["italic"] = request.italic
        if request.underline:
            font_kwargs["underline"] = request.underline
        if request.color:
            font_kwargs["color"] = request.color

        font = Font(**font_kwargs)

        # Apply to range
        for row in ws[request.range_ref]:
            for cell in row if isinstance(row, tuple) else [row]:
                cell.font = font

        # Save workbook
        wb.save(request.workbook_path)
        wb.close()

        return OperationResult(success=True, message=f"Font formatting applied to {request.range_ref}")

    except Exception as e:
        return OperationResult(success=False, message=f"Failed to apply font formatting: {str(e)}")


def format_fill(request: FillFormatRequest) -> OperationResult:
    """
    Apply background fill (color) to a range of cells.

    Args:
        request: FillFormatRequest with formatting parameters

    Returns:
        OperationResult with success status
    """
    try:
        # Validate file path
        is_valid, error = validate_file_path(request.workbook_path, must_exist=True)
        if not is_valid:
            return OperationResult(success=False, message=error)

        # Validate range
        is_valid, error = validate_range_reference(request.range_ref)
        if not is_valid:
            return OperationResult(success=False, message=error)

        # Load workbook
        wb = load_workbook(request.workbook_path)

        # Check if sheet exists
        if request.sheet_name not in wb.sheetnames:
            wb.close()
            return OperationResult(
                success=False,
                message=f"Sheet '{request.sheet_name}' not found. Available sheets: {wb.sheetnames}",
            )

        ws = wb[request.sheet_name]

        # Create fill object
        fill = PatternFill(start_color=request.color, end_color=request.color, fill_type=cast(Any, request.fill_type))

        # Apply to range
        for row in ws[request.range_ref]:
            for cell in row if isinstance(row, tuple) else [row]:
                cell.fill = fill

        # Save workbook
        wb.save(request.workbook_path)
        wb.close()

        return OperationResult(success=True, message=f"Fill formatting applied to {request.range_ref}")

    except Exception as e:
        return OperationResult(success=False, message=f"Failed to apply fill formatting: {str(e)}")


def format_border(request: BorderFormatRequest) -> OperationResult:
    """
    Apply border formatting to a range of cells.

    Args:
        request: BorderFormatRequest with formatting parameters

    Returns:
        OperationResult with success status
    """
    try:
        # Validate file path
        is_valid, error = validate_file_path(request.workbook_path, must_exist=True)
        if not is_valid:
            return OperationResult(success=False, message=error)

        # Validate range
        is_valid, error = validate_range_reference(request.range_ref)
        if not is_valid:
            return OperationResult(success=False, message=error)

        # Load workbook
        wb = load_workbook(request.workbook_path)

        # Check if sheet exists
        if request.sheet_name not in wb.sheetnames:
            wb.close()
            return OperationResult(
                success=False,
                message=f"Sheet '{request.sheet_name}' not found. Available sheets: {wb.sheetnames}",
            )

        ws = wb[request.sheet_name]

        # Create side object for border
        side_kwargs: dict[str, Any] = {"style": cast(Any, request.style)}
        if request.color:
            side_kwargs["color"] = request.color

        side = Side(**side_kwargs)

        # Create border object with specified sides
        border_kwargs = {}
        if "top" in request.sides:
            border_kwargs["top"] = side
        if "bottom" in request.sides:
            border_kwargs["bottom"] = side
        if "left" in request.sides:
            border_kwargs["left"] = side
        if "right" in request.sides:
            border_kwargs["right"] = side

        border = Border(**border_kwargs)

        # Apply to range
        for row in ws[request.range_ref]:
            for cell in row if isinstance(row, tuple) else [row]:
                cell.border = border

        # Save workbook
        wb.save(request.workbook_path)
        wb.close()

        return OperationResult(success=True, message=f"Border formatting applied to {request.range_ref}")

    except Exception as e:
        return OperationResult(success=False, message=f"Failed to apply border formatting: {str(e)}")


def format_alignment(request: AlignmentFormatRequest) -> OperationResult:
    """
    Apply alignment formatting to a range of cells.

    Args:
        request: AlignmentFormatRequest with formatting parameters

    Returns:
        OperationResult with success status
    """
    try:
        # Validate file path
        is_valid, error = validate_file_path(request.workbook_path, must_exist=True)
        if not is_valid:
            return OperationResult(success=False, message=error)

        # Validate range
        is_valid, error = validate_range_reference(request.range_ref)
        if not is_valid:
            return OperationResult(success=False, message=error)

        # Load workbook
        wb = load_workbook(request.workbook_path)

        # Check if sheet exists
        if request.sheet_name not in wb.sheetnames:
            wb.close()
            return OperationResult(
                success=False,
                message=f"Sheet '{request.sheet_name}' not found. Available sheets: {wb.sheetnames}",
            )

        ws = wb[request.sheet_name]

        # Create alignment object
        alignment_kwargs = {}
        if request.horizontal:
            alignment_kwargs["horizontal"] = request.horizontal
        if request.vertical:
            alignment_kwargs["vertical"] = request.vertical
        if request.wrap_text is not None:
            alignment_kwargs["wrap_text"] = request.wrap_text
        if request.text_rotation is not None:
            alignment_kwargs["text_rotation"] = request.text_rotation

        alignment = Alignment(**alignment_kwargs)

        # Apply to range
        for row in ws[request.range_ref]:
            for cell in row if isinstance(row, tuple) else [row]:
                cell.alignment = alignment

        # Save workbook
        wb.save(request.workbook_path)
        wb.close()

        return OperationResult(success=True, message=f"Alignment formatting applied to {request.range_ref}")

    except Exception as e:
        return OperationResult(success=False, message=f"Failed to apply alignment formatting: {str(e)}")


def format_number(request: NumberFormatRequest) -> OperationResult:
    """
    Apply number formatting to a range of cells.

    Args:
        request: NumberFormatRequest with format string

    Returns:
        OperationResult with success status
    """
    try:
        # Validate file path
        is_valid, error = validate_file_path(request.workbook_path, must_exist=True)
        if not is_valid:
            return OperationResult(success=False, message=error)

        # Validate range
        is_valid, error = validate_range_reference(request.range_ref)
        if not is_valid:
            return OperationResult(success=False, message=error)

        # Load workbook
        wb = load_workbook(request.workbook_path)

        # Check if sheet exists
        if request.sheet_name not in wb.sheetnames:
            wb.close()
            return OperationResult(
                success=False,
                message=f"Sheet '{request.sheet_name}' not found. Available sheets: {wb.sheetnames}",
            )

        ws = wb[request.sheet_name]

        # Apply number format to range
        for row in ws[request.range_ref]:
            for cell in row if isinstance(row, tuple) else [row]:
                cell.number_format = request.format_string

        # Save workbook
        wb.save(request.workbook_path)
        wb.close()

        return OperationResult(success=True, message=f"Number formatting applied to {request.range_ref}")

    except Exception as e:
        return OperationResult(success=False, message=f"Failed to apply number formatting: {str(e)}")
