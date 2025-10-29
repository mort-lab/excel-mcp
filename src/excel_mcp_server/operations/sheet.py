"""Sheet operations"""


from openpyxl import load_workbook

from ..models import SheetCreateRequest, SheetRenameRequest, SheetResult
from ..utils.validators import validate_file_path, validate_sheet_name


def create(request: SheetCreateRequest) -> SheetResult:
    """
    Create a new worksheet in the workbook.

    Args:
        request: SheetCreateRequest with workbook path, sheet name, and optional index

    Returns:
        SheetResult with success status
    """
    try:
        # Validate file path
        is_valid, error = validate_file_path(request.workbook_path, must_exist=True)
        if not is_valid:
            return SheetResult(success=False, message=error)

        # Validate sheet name
        is_valid, error = validate_sheet_name(request.sheet_name)
        if not is_valid:
            return SheetResult(success=False, message=error)

        # Load workbook
        wb = load_workbook(request.workbook_path)

        # Check if sheet already exists
        if request.sheet_name in wb.sheetnames:
            wb.close()
            return SheetResult(
                success=False,
                message=f"Sheet '{request.sheet_name}' already exists",
                sheet_name=request.sheet_name,
            )

        # Create new sheet
        if request.index is not None:
            wb.create_sheet(request.sheet_name, request.index)
        else:
            wb.create_sheet(request.sheet_name)

        # Save workbook
        wb.save(request.workbook_path)
        wb.close()

        return SheetResult(
            success=True, message=f"Sheet '{request.sheet_name}' created successfully", sheet_name=request.sheet_name
        )

    except Exception as e:
        return SheetResult(success=False, message=f"Failed to create sheet: {str(e)}")


def delete(workbook_path: str, sheet_name: str) -> SheetResult:
    """
    Delete a worksheet from the workbook.

    Args:
        workbook_path: Path to the Excel workbook
        sheet_name: Name of the sheet to delete

    Returns:
        SheetResult with success status
    """
    try:
        # Validate file path
        is_valid, error = validate_file_path(workbook_path, must_exist=True)
        if not is_valid:
            return SheetResult(success=False, message=error)

        # Load workbook
        wb = load_workbook(workbook_path)

        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            wb.close()
            return SheetResult(
                success=False,
                message=f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}",
                sheet_name=sheet_name,
            )

        # Prevent deleting the last sheet
        if len(wb.sheetnames) == 1:
            wb.close()
            return SheetResult(success=False, message="Cannot delete the last sheet in the workbook")

        # Delete sheet
        del wb[sheet_name]

        # Save workbook
        wb.save(workbook_path)
        wb.close()

        return SheetResult(success=True, message=f"Sheet '{sheet_name}' deleted successfully", sheet_name=sheet_name)

    except Exception as e:
        return SheetResult(success=False, message=f"Failed to delete sheet: {str(e)}")


def rename(request: SheetRenameRequest) -> SheetResult:
    """
    Rename a worksheet.

    Args:
        request: SheetRenameRequest with workbook path, old name, and new name

    Returns:
        SheetResult with success status
    """
    try:
        # Validate file path
        is_valid, error = validate_file_path(request.workbook_path, must_exist=True)
        if not is_valid:
            return SheetResult(success=False, message=error)

        # Validate new sheet name
        is_valid, error = validate_sheet_name(request.new_name)
        if not is_valid:
            return SheetResult(success=False, message=error)

        # Load workbook
        wb = load_workbook(request.workbook_path)

        # Check if old sheet exists
        if request.old_name not in wb.sheetnames:
            wb.close()
            return SheetResult(
                success=False,
                message=f"Sheet '{request.old_name}' not found. Available sheets: {wb.sheetnames}",
                sheet_name=request.old_name,
            )

        # Check if new name already exists
        if request.new_name in wb.sheetnames:
            wb.close()
            return SheetResult(success=False, message=f"Sheet '{request.new_name}' already exists")

        # Rename sheet
        ws = wb[request.old_name]
        ws.title = request.new_name

        # Save workbook
        wb.save(request.workbook_path)
        wb.close()

        return SheetResult(
            success=True,
            message=f"Sheet renamed from '{request.old_name}' to '{request.new_name}'",
            sheet_name=request.new_name,
        )

    except Exception as e:
        return SheetResult(success=False, message=f"Failed to rename sheet: {str(e)}")


def copy_sheet(workbook_path: str, source_sheet: str, new_name: str) -> SheetResult:
    """
    Copy a worksheet within the workbook.

    Args:
        workbook_path: Path to the Excel workbook
        source_sheet: Name of the sheet to copy
        new_name: Name for the new sheet

    Returns:
        SheetResult with success status
    """
    try:
        # Validate file path
        is_valid, error = validate_file_path(workbook_path, must_exist=True)
        if not is_valid:
            return SheetResult(success=False, message=error)

        # Validate new sheet name
        is_valid, error = validate_sheet_name(new_name)
        if not is_valid:
            return SheetResult(success=False, message=error)

        # Load workbook
        wb = load_workbook(workbook_path)

        # Check if source sheet exists
        if source_sheet not in wb.sheetnames:
            wb.close()
            return SheetResult(
                success=False,
                message=f"Sheet '{source_sheet}' not found. Available sheets: {wb.sheetnames}",
                sheet_name=source_sheet,
            )

        # Check if new name already exists
        if new_name in wb.sheetnames:
            wb.close()
            return SheetResult(success=False, message=f"Sheet '{new_name}' already exists")

        # Copy sheet
        source = wb[source_sheet]
        target = wb.copy_worksheet(source)
        target.title = new_name

        # Save workbook
        wb.save(workbook_path)
        wb.close()

        return SheetResult(
            success=True, message=f"Sheet '{source_sheet}' copied to '{new_name}'", sheet_name=new_name
        )

    except Exception as e:
        return SheetResult(success=False, message=f"Failed to copy sheet: {str(e)}")
