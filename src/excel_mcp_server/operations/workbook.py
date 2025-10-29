"""Workbook operations"""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from ..models import WorkbookInfo, WorkbookResult
from ..utils.validators import validate_file_path


def create(file_path: str) -> WorkbookResult:
    """
    Create a new Excel workbook.

    Args:
        file_path: Path where the workbook will be created

    Returns:
        WorkbookResult with success status
    """
    try:
        # Validate file path
        is_valid, error = validate_file_path(file_path, must_exist=False)
        if not is_valid:
            return WorkbookResult(success=False, message=error)

        # Check if file already exists
        if Path(file_path).exists():
            return WorkbookResult(
                success=False, message=f"File already exists: {file_path}", file_path=file_path
            )

        # Create new workbook
        wb = Workbook()
        wb.save(file_path)

        return WorkbookResult(success=True, message="Workbook created successfully", file_path=file_path)

    except Exception as e:
        return WorkbookResult(success=False, message=f"Failed to create workbook: {str(e)}")


def open_file(file_path: str) -> WorkbookInfo:
    """
    Open an existing workbook and return metadata.

    Args:
        file_path: Path to the Excel workbook

    Returns:
        WorkbookInfo with workbook metadata
    """
    try:
        # Validate file path
        is_valid, error = validate_file_path(file_path, must_exist=True)
        if not is_valid:
            raise ValueError(error)

        # Load workbook
        wb = load_workbook(file_path, read_only=True, data_only=False)

        # Get sheet names
        sheets = wb.sheetnames
        sheet_count = len(sheets)

        # Get file size
        file_size = Path(file_path).stat().st_size

        wb.close()

        return WorkbookInfo(
            file_path=file_path, sheets=sheets, sheet_count=sheet_count, file_size=file_size
        )

    except FileNotFoundError as e:
        raise ValueError(f"File not found: {file_path}") from e
    except InvalidFileException as e:
        raise ValueError(f"Invalid Excel file: {file_path}") from e
    except Exception as e:
        raise ValueError(f"Failed to open workbook: {str(e)}") from e


def list_sheets(file_path: str) -> list[str]:
    """
    List all sheet names in a workbook.

    Args:
        file_path: Path to the Excel workbook

    Returns:
        List of sheet names
    """
    try:
        # Validate file path
        is_valid, error = validate_file_path(file_path, must_exist=True)
        if not is_valid:
            raise ValueError(error)

        # Load workbook
        wb = load_workbook(file_path, read_only=True, keep_links=False)
        sheets = wb.sheetnames
        wb.close()

        return sheets

    except FileNotFoundError as e:
        raise ValueError(f"File not found: {file_path}") from e
    except Exception as e:
        raise ValueError(f"Failed to list sheets: {str(e)}") from e


def save(file_path: str) -> WorkbookResult:
    """
    Save changes to a workbook.

    Args:
        file_path: Path to the Excel workbook

    Returns:
        WorkbookResult with success status
    """
    try:
        # Validate file path
        is_valid, error = validate_file_path(file_path, must_exist=True)
        if not is_valid:
            return WorkbookResult(success=False, message=error)

        # Load and save workbook (this validates it's a proper Excel file)
        wb = load_workbook(file_path)
        wb.save(file_path)
        wb.close()

        return WorkbookResult(success=True, message="Workbook saved successfully", file_path=file_path)

    except FileNotFoundError:
        return WorkbookResult(success=False, message=f"File not found: {file_path}")
    except Exception as e:
        return WorkbookResult(success=False, message=f"Failed to save workbook: {str(e)}")


def get_info(file_path: str) -> dict:
    """
    Get comprehensive workbook information.

    Args:
        file_path: Path to the Excel workbook

    Returns:
        Dictionary with workbook metadata
    """
    try:
        info = open_file(file_path)
        return info.model_dump()
    except Exception as e:
        return {"error": str(e), "file_path": file_path}
