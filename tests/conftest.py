"""Shared pytest fixtures"""

import pytest
from pathlib import Path
import tempfile
from openpyxl import Workbook


@pytest.fixture
def temp_excel_file():
    """Provides a temporary Excel file path that will be cleaned up after use"""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        file_path = Path(tmp.name)
    yield str(file_path)
    # Cleanup
    if file_path.exists():
        file_path.unlink()


@pytest.fixture
def sample_workbook(temp_excel_file):
    """Creates a sample workbook with data for testing"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Add some sample data
    ws["A1"] = "Name"
    ws["B1"] = "Age"
    ws["C1"] = "City"

    ws["A2"] = "Alice"
    ws["B2"] = 30
    ws["C2"] = "New York"

    ws["A3"] = "Bob"
    ws["B3"] = 25
    ws["C3"] = "Los Angeles"

    ws["A4"] = "Charlie"
    ws["B4"] = 35
    ws["C4"] = "Chicago"

    wb.save(temp_excel_file)
    wb.close()

    yield temp_excel_file


@pytest.fixture
def temp_dir():
    """Provides a temporary directory for file operations"""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)
